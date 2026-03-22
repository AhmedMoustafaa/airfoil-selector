"""
AirfoilSelector
===============
Multi-objective airfoil selector for UAV design.

Aerodynamic metrics supported
-----------------------------------------
  cl          — maximum lift coefficient          (takeoff / high-lift)
  cl_cd       — CL / CD                           (glide ratio / jet range)
  cl32_cd     — CL^(3/2) / CD                    (propeller endurance, Breguet)
  cl12_cd     — CL^(1/2) / CD                    (minimum power / prop range)
  min_cd      — minimum drag coefficient          (cruise drag budget)

Constraint types  (used with apply_constraints)
-------------------------------------------------
  thickness   — airfoil t/c ratio bounds
  camber      — camber ratio bounds
  cl_max      — minimum required CL_max at a given Re
  cm          — maximum allowed Cm  (cm_max=0.0 removes reflexed airfoils)
"""

import warnings
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import aerosandbox as asb
import os
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Module-level metric functions
# Each f(cl_curve, cd_curve) -> metric curve of the same shape.
# They are pure functions so they can be used both inside the AOA-sweep loop
# and directly in plotting / reporting without going through the class.
# ---------------------------------------------------------------------------

def _metric_cl(cl, cd):
    return np.asarray(cl, dtype=float)

def _metric_cl_cd(cl, cd):
    cl, cd = np.asarray(cl, dtype=float), np.asarray(cd, dtype=float)
    with np.errstate(divide='ignore', invalid='ignore'):
        return np.where(cd > 0, cl / cd, 0.0)

def _metric_cl32_cd(cl, cd):
    """CL^(3/2) / CD — maximise for best propeller endurance (Breguet prop endurance)."""
    cl, cd = np.asarray(cl, dtype=float), np.asarray(cd, dtype=float)
    with np.errstate(divide='ignore', invalid='ignore'):
        return np.where((cl > 0) & (cd > 0), cl ** 1.5 / cd, 0.0)

def _metric_cl12_cd(cl, cd):
    """CL^(1/2) / CD — maximise for minimum-power speed / best propeller range."""
    cl, cd = np.asarray(cl, dtype=float), np.asarray(cd, dtype=float)
    with np.errstate(divide='ignore', invalid='ignore'):
        return np.where((cl > 0) & (cd > 0), cl ** 0.5 / cd, 0.0)

# Registry: metric name -> (fn, higher_is_better)
_OBJECTIVE_REGISTRY = {
    'cl':      (_metric_cl,      True),
    'cl_cd':   (_metric_cl_cd,   True),
    'cl32_cd': (_metric_cl32_cd, True),
    'cl12_cd': (_metric_cl12_cd, True),
    # 'min_cd' has a different loop structure (CL interpolation) — handled separately
}

_OBJECTIVE_LABELS = {
    'cl':      'Max CL  (high-lift)',
    'cl_cd':   'Max CL/CD  (glide / jet range)',
    'cl32_cd': 'Max CL^1.5/CD  (prop endurance)',
    'cl12_cd': 'Max CL^0.5/CD  (prop range / min-power)',
    'min_cd':  'Min CD  (cruise drag budget)',
}

class AirfoilSelector:

    def __init__(self, db_path='full_data.pkl'):
        self.df = pd.read_pickle(db_path)

        self.cl        = self.df['cl'].to_numpy()
        self.cd        = self.df['cd'].to_numpy()
        self.cm        = self.df['cm'].to_numpy()
        self.cl_cd     = self.df['cl_cd'].to_numpy()
        self.names     = self.df['name'].to_numpy()
        self.thickness = self.df['thickness'].to_numpy()
        self.camber    = self.df['camber'].to_numpy()

        # Re and alpha are read directly from the database so the selector
        # automatically adapts when the database is rebuilt with a different
        # Re range or AOA grid — no code changes required.
        if 're' in self.df.columns and 'alpha' in self.df.columns:
            self.Re    = np.array(self.df['re'].iloc[0])
            self.alpha = np.array(self.df['alpha'].iloc[0])
        else:
            # Legacy database (no re/alpha columns): infer from data shape.
            # Shape of cl per airfoil is (n_re, n_alpha).
            sample = np.array(self.cl[0])
            n_re, n_alpha = sample.shape
            warnings.warn(
                f"Database does not contain 're' and 'alpha' columns. "
                f"Inferring grid: {n_re} Re points, {n_alpha} alpha points. "
                f"Rebuild the database with the updated generation script to "
                f"store the grid explicitly.",
                stacklevel=2,
            )
            # Fall back to the shapes that were standard before the update
            self.Re    = np.linspace(1e5, 1e6, n_re)
            self.alpha = np.linspace(-5, 20, n_alpha)

    @staticmethod
    def distribute(center: float, w1: float, w2: float, width: int,
                   arr: np.ndarray) -> np.ndarray:
        """
        Build a linear-ramp weight array peaked at `center`.

        Parameters
        ----------
        center : float  – value in arr where peak weight w1 is applied
        w1     : float  – peak weight (must be >= w2)
        w2     : float  – baseline weight outside the ramp region
        width  : int    – half-width of the ramp in index units
        arr    : ndarray – target array (e.g. self.alpha or self.Re)
        """
        if w1 < w2:
            raise ValueError(f"w1 ({w1}) must be >= w2 ({w2}).")
        if width <= 0:
            raise ValueError("width must be a positive integer.")

        dy_dx    = (w1 - w2) / width
        weights  = np.ones(arr.shape) * w2
        ci       = int(np.searchsorted(arr, center))
        min_i    = ci - width

        for i in range(len(arr)):
            if abs(i - ci) <= width:
                weights[i] = (w2 + dy_dx * (i - min_i) if i <= ci
                              else w1 - dy_dx * (i - ci))
        return weights

    @staticmethod
    def weighted_mean(arr, weights) -> float:
        a, w = np.ravel(arr), np.ravel(weights)
        n = min(len(a), len(w))
        if len(a) != len(w):
            warnings.warn(
                f"weighted_mean: length mismatch ({len(a)} vs {len(w)}), "
                f"truncating to {n}.", stacklevel=2)
        return float(np.average(a[:n], weights=w[:n]))

    def _re_index(self, re_val) -> int:
        """
        Index into self.Re for re_val.

        Raises a clear ValueError if the value is missing, with a hint when
        the value is out of the database bounds (so the user knows whether
        to rebuild the database or pick a different Re).
        """
        idx = np.where(np.isclose(self.Re, re_val))[0]
        if not len(idx):
            lo, hi = self.Re.min(), self.Re.max()
            hint = ""
            if re_val < lo:
                hint = f" (below database minimum Re={lo:.2e})"
            elif re_val > hi:
                hint = f" (above database maximum Re={hi:.2e})"
            raise ValueError(
                f"Re={re_val:.2e} not in database{hint}. "
                f"Available: {self.Re}")
        return int(idx[0])

    def _re_in_range(self, re_val, Re_range) -> bool:
        return bool(any(np.isclose(re_val, Re_range)))

    def _validate_re_range(self, Re_range):
        """
        Warn if any requested Re values are outside the database bounds or
        are not present as grid points — called at the start of every objective
        method so the user gets a readable message before the loop runs.
        """
        db_min, db_max = self.Re.min(), self.Re.max()
        missing = []
        for rv in Re_range:
            if not any(np.isclose(rv, self.Re)):
                if rv < db_min or rv > db_max:
                    missing.append(
                        f"{rv:.2e} (out of database range "
                        f"[{db_min:.2e}, {db_max:.2e}])")
                else:
                    nearest = self.Re[np.argmin(np.abs(self.Re - rv))]
                    missing.append(
                        f"{rv:.2e} (not a grid point; nearest: {nearest:.2e})")
        if missing:
            warnings.warn(
                f"The following Re values are not in the database and will be "
                f"skipped: {', '.join(missing)}. "
                f"Available Re grid: {self.Re}",
                stacklevel=3,
            )

    def _validate_aoa_range(self, aoa_weights):
        """
        Raise if aoa_weights length does not match the database alpha grid.
        Catches the common mistake of building weights for a fixed 26-point grid
        after the database is rebuilt with a different alpha resolution.
        """
        if len(aoa_weights) != len(self.alpha):
            raise ValueError(
                f"aoa_weights length ({len(aoa_weights)}) does not match "
                f"the database alpha grid ({len(self.alpha)} points, "
                f"{self.alpha[0]:.1f}° to {self.alpha[-1]:.1f}°). "
                f"Rebuild aoa_weights using selector.alpha as the reference array."
            )

    def _airfoil_index(self, name) -> int:
        idx = np.where(self.names == name)[0]
        if not len(idx):
            raise KeyError(f"Airfoil '{name}' not found in database.")
        return int(idx[0])

    def _sort_dict(self, d, reverse=True):
        return dict(sorted(d.items(), key=lambda x: x[1], reverse=reverse))

    def _drop_zeros(self, d):
        return {k: v for k, v in d.items() if v != 0}

    def get_top_n(self, result_dict, n=5, ascending=False):
        """
        Return the top-n entries from a scored dict.

        ascending=False (default) -> highest scores first (CL, L/D objectives).
        ascending=True            -> lowest scores first  (CD objective).
        """
        return dict(
            list(self._sort_dict(result_dict, reverse=not ascending).items())[:n])

    def _cl_spans_range(self, cl_curve, cl_range, tol=0.03) -> bool:
        """True if cl_curve covers the full cl_range within tolerance."""
        return (np.min(cl_curve) - tol <= cl_range[0] and
                np.max(cl_curve) >= cl_range[-1] - tol)

    def _cl_within_aoa_range(self, cl_range, cl_curve, aoa_range) -> bool:
        if not self._cl_spans_range(cl_curve, cl_range):
            return False
        cl_arr = np.array(cl_curve)
        lo = np.searchsorted(self.alpha, aoa_range[0])
        hi = np.searchsorted(self.alpha, aoa_range[-1])
        idxs = np.clip(np.searchsorted(cl_arr, cl_range), 0, len(cl_arr) - 1)
        return lo <= idxs[0] and idxs[-1] <= hi

    def _interp_cd_at_cl(self, cl_values, cl_curve, cd_curve):
        cl_a, cd_a = np.array(cl_curve), np.array(cd_curve)
        s = np.argsort(cl_a)
        return np.interp(cl_values, cl_a[s], cd_a[s])

    def _aoa_sweep(self, metric_fn, Re_range, re_weights, aoa_weights):
        """
        Core loop shared by all AOA-sweep objectives.

        For each airfoil:
          1. For each Re in Re_range, compute metric_fn(cl_curve, cd_curve).
          2. Weighted-mean over AOA  -> scalar per Re.
          3. Weighted-mean over Re   -> final score.
        """
        self._validate_re_range(Re_range)
        self._validate_aoa_range(aoa_weights)
        scores = {}
        for i, name in enumerate(self.names):
            re_scores = np.zeros(len(Re_range))
            j2 = 0
            for j in range(len(self.Re)):
                if self._re_in_range(self.Re[j], Re_range):
                    curve = metric_fn(self.cl[i][j], self.cd[i][j])
                    re_scores[j2] = self.weighted_mean(curve, aoa_weights)
                    j2 += 1
            scores[name] = self.weighted_mean(re_scores, re_weights)
        return self._drop_zeros(scores)

    def _min_cd_loop(self, Re_range, re_weights, cl_range, cl_weights, aoa_range):
        self._validate_re_range(Re_range)
        scores = {}
        for i, name in enumerate(self.names):
            re_scores = np.zeros(len(Re_range))
            j2 = 0
            for j in range(len(self.Re)):
                if self._re_in_range(self.Re[j], Re_range):
                    if self._cl_within_aoa_range(cl_range, self.cl[i][j], aoa_range):
                        cd_at_cl = self._interp_cd_at_cl(
                            cl_range, self.cl[i][j], self.cd[i][j])
                        re_scores[j2] = self.weighted_mean(cd_at_cl, cl_weights)
                        j2 += 1
            scores[name] = self.weighted_mean(re_scores, re_weights)
        return self._drop_zeros(scores)

    def compute_objective(self, metric, Re_range, re_weights, aoa_weights=None,
                          cl_range=None, cl_weights=None, aoa_range=None):
        """
        Unified objective dispatcher.

        Parameters
        ----------
        metric      : str  — 'cl' | 'cl_cd' | 'cl32_cd' | 'cl12_cd' | 'min_cd'
        Re_range    : list[float]
        re_weights  : list[float]
        aoa_weights : ndarray  — required for all except 'min_cd'
        cl_range    : ndarray  — required for 'min_cd'
        cl_weights  : ndarray  — required for 'min_cd'
        aoa_range   : ndarray  — required for 'min_cd'

        Returns
        -------
        dict  {airfoil_name: score}

        Examples
        --------
        >>> # Endurance (prop)
        >>> scores = selector.compute_objective(
        ...     'cl32_cd', Re_range, Re_weights, aoa_weights)
        >>> # Cruise drag
        >>> scores = selector.compute_objective(
        ...     'min_cd', Re_range, Re_weights,
        ...     cl_range=cl_range, cl_weights=cl_weights, aoa_range=aoa_range)
        """
        if metric in _OBJECTIVE_REGISTRY:
            if aoa_weights is None:
                raise ValueError(f"aoa_weights is required for metric='{metric}'.")
            fn, _ = _OBJECTIVE_REGISTRY[metric]
            return self._aoa_sweep(fn, Re_range, re_weights, aoa_weights)

        if metric == 'min_cd':
            if any(v is None for v in (cl_range, cl_weights, aoa_range)):
                raise ValueError(
                    "cl_range, cl_weights, and aoa_range are required for metric='min_cd'.")
            return self._min_cd_loop(Re_range, re_weights, cl_range, cl_weights, aoa_range)

        raise ValueError(
            f"Unknown metric '{metric}'. "
            f"Valid: {list(_OBJECTIVE_REGISTRY) + ['min_cd']}")

    def max_cl(self, Re_range, re_weights, aoa_weights):
        """Maximise CL — best for takeoff / high-lift selection."""
        return self._aoa_sweep(_metric_cl, Re_range, re_weights, aoa_weights)

    def max_cl_cd(self, Re_range, re_weights, aoa_weights):
        """Maximise CL/CD — best glide ratio / jet range."""
        return self._aoa_sweep(_metric_cl_cd, Re_range, re_weights, aoa_weights)

    def max_cl32_cd(self, Re_range, re_weights, aoa_weights):
        """Maximise CL^(3/2)/CD — best propeller endurance (Breguet prop)."""
        return self._aoa_sweep(_metric_cl32_cd, Re_range, re_weights, aoa_weights)

    def max_cl12_cd(self, Re_range, re_weights, aoa_weights):
        """Maximise CL^(1/2)/CD — minimum power speed / best propeller range."""
        return self._aoa_sweep(_metric_cl12_cd, Re_range, re_weights, aoa_weights)

    def min_cd(self, Re_range, re_weights, cl_range, cl_weights, aoa_range):
        """Minimise CD across a target CL range and AOA band."""
        return self._min_cd_loop(Re_range, re_weights, cl_range, cl_weights, aoa_range)

    def constraint_thickness(self, airfoil_dict, thick_min=0.0, thick_max=1.0):
        """Keep airfoils where  thick_min ≤ t/c ≤ thick_max  (fractions, e.g. 0.12)."""
        d = airfoil_dict.copy()
        for name in list(d):
            idx = np.where(self.names == name)[0]
            if not len(idx) or not (thick_min <= self.thickness[idx[0]] <= thick_max):
                d.pop(name)
        return d

    def constraint_camber(self, airfoil_dict, camb_min=0.0, camb_max=1.0):
        """Keep airfoils where  camb_min ≤ camber ≤ camb_max  (fractions)."""
        d = airfoil_dict.copy()
        for name in list(d):
            idx = np.where(self.names == name)[0]
            if not len(idx) or not (camb_min <= self.camber[idx[0]] <= camb_max):
                d.pop(name)
        return d

    def constraint_cl_max(self, airfoil_dict, cl_min, re_target):
        """
        Keep airfoils whose peak CL at re_target is >= cl_min.

        Tip: pass  cl_min = CL_3D_required * (1.10-1.2)  to account for finite-span
        correction from 2-D to 3-D.
        """
        re_idx = self._re_index(re_target)
        d = airfoil_dict.copy()
        for name in list(d):
            idx = np.where(self.names == name)[0]
            if not len(idx):
                d.pop(name); continue
            if float(np.max(np.asarray(self.cl[idx[0]][re_idx], dtype=float))) < cl_min:
                d.pop(name)
        return d

    def constraint_cm(self, airfoil_dict, re_target, cm_max=0.0):
        """
        Keep airfoils whose maximum Cm at re_target is <= cm_max.

        Reflexed airfoils (flying-wing / tailless designs) have Cm > 0 over part
        or all of their AOA range. Setting cm_max=0.0 (default) removes all such
        sections, retaining only conventional cambered and symmetric airfoils.

        Guidance on cm_max
        -------------------
        cm_max =  0.00  — exclude all reflexed sections (standard tailed UAV)
        cm_max =  0.02  — allow mildly reflexed sections (semi-tailless)
        cm_max = -0.02  — require a stable nose-down moment (highly stable designs)
        """
        re_idx = self._re_index(re_target)
        d = airfoil_dict.copy()
        for name in list(d):
            idx = np.where(self.names == name)[0]
            if not len(idx):
                d.pop(name); continue
            cm_curve = np.asarray(self.cm[idx[0]][re_idx], dtype=float)
            if float(np.max(cm_curve)) > cm_max:
                d.pop(name)
        return d

    def apply_constraints(self, airfoil_dict, constraints):
        """
        Apply an ordered chain of constraints in a single call.

        Parameters
        ----------
        airfoil_dict : dict  {name: score}
        constraints  : list[dict]
            Supported constraint specs::

                {'type': 'thickness', 'min': 0.10, 'max': 0.20}
                {'type': 'camber',    'min': 0.00, 'max': 0.05}
                {'type': 'cl_max',    'min': 1.30, 're':  3e5}
                {'type': 'cm',        'max': 0.00, 're':  3e5}

        Returns
        -------
        dict  {name: score}  — filtered copy

        Example
        -------
        >>> filtered = selector.apply_constraints(raw_dict, [
        ...     {'type': 'thickness', 'min': 0.10, 'max': 0.20},
        ...     {'type': 'camber',    'min': 0.00, 'max': 0.05},
        ...     {'type': 'cl_max',    'min': 1.30, 're':  3e5},
        ...     {'type': 'cm',        'max': 0.00, 're':  3e5},
        ... ])
        """
        _HANDLERS = {
            'thickness': lambda d, s: self.constraint_thickness(
                d, thick_min=s.get('min', 0.0), thick_max=s.get('max', 1.0)),
            'camber':    lambda d, s: self.constraint_camber(
                d, camb_min=s.get('min', 0.0),  camb_max=s.get('max', 1.0)),
            'cl_max':    lambda d, s: self.constraint_cl_max(
                d, cl_min=s['min'], re_target=s['re']),
            'cm':        lambda d, s: self.constraint_cm(
                d, re_target=s['re'], cm_max=s.get('max', 0.0)),
        }

        d = airfoil_dict.copy()
        before = len(d)

        for spec in constraints:
            ctype = spec.get('type')
            if ctype not in _HANDLERS:
                raise ValueError(
                    f"Unknown constraint type '{ctype}'. "
                    f"Valid types: {list(_HANDLERS)}")

            # Validate required keys before calling
            required = {'cl_max': ('min', 're'), 'cm': ('re',)}.get(ctype, ())
            missing = [k for k in required if k not in spec]
            if missing:
                raise ValueError(
                    f"Constraint '{ctype}' is missing required key(s): {missing}")

            d = _HANDLERS[ctype](d, spec)
            after = len(d)

            if after == 0:
                warnings.warn(
                    f"apply_constraints: no airfoils remain after "
                    f"'{ctype}' constraint (started with {before}).",
                    stacklevel=2)
                return d
            before = after

        return d

    CRITERIA_PRESETS = {
        'max_cl':     {'cl': 1.0},
        'min_cd':     {'cd': 1.0},
        'max_cl_cd':  {'cl_cd':   1.0},
        'max_cl32_cd':{'cl32_cd': 1.0},
        'max_cl12_cd':{'cl12_cd': 1.0},
    }

    _SCORE_METRICS_HIB = {
        # key: higher_is_better
        'cl':          True,
        'cd':          False,
        'cm':          False,   # |Cm| — lower is better (more stable / less pitchy)
        'cl_cd':       True,
        'cl32_cd':     True,
        'cl12_cd':     True,
        'alpha_stall': True,
    }

    def score_airfoils(self, airfoil_names, re_target,
                       score_weights=None, criteria='custom'):
        """
        Rank candidates with a normalised multi-criteria weighted score.

        All metrics are min-max normalised to [0, 1] across the candidate set
        before weighting — so weights are physically comparable regardless of
        scale.

        Parameters
        ----------
        airfoil_names : list[str]
        re_target     : float
        score_weights : dict, optional
            Supported keys::

                'cl'          — weight for max CL          (higher is better)
                'cd'          — weight for min CD          (lower  is better)
                'cm'          — weight for min |Cm|        (lower  is better)
                'cl_cd'       — weight for max CL/CD       (higher is better)
                'cl32_cd'     — weight for max CL^1.5/CD   (higher is better)
                'cl12_cd'     — weight for max CL^0.5/CD   (higher is better)
                'alpha_stall' — weight for max stall angle  (higher is better)

            Weights do not need to sum to 1.
        criteria : str
            One of the CRITERIA_PRESETS keys, or 'custom'.

        Returns
        -------
        dict  {name: score}  sorted highest -> lowest.
        """
        if criteria in self.CRITERIA_PRESETS:
            weights = self.CRITERIA_PRESETS[criteria]
        elif criteria == 'custom':
            if score_weights is None:
                raise ValueError("Provide score_weights when criteria='custom'.")
            weights = score_weights
        else:
            raise ValueError(
                f"Unknown criteria '{criteria}'. "
                f"Valid: {list(self.CRITERIA_PRESETS) + ['custom']}")

        re_idx = self._re_index(re_target)

        raw = {}
        for name in airfoil_names:
            try:
                idx = self._airfoil_index(name)
            except KeyError as e:
                warnings.warn(str(e), stacklevel=2); continue

            cl_c = np.asarray(self.cl[idx][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[idx][re_idx], dtype=float)
            cm_c = np.asarray(self.cm[idx][re_idx], dtype=float)

            raw[name] = {
                'cl':          float(np.max(cl_c)),
                'cd':          float(np.min(cd_c)),
                'cm':          float(np.min(np.abs(cm_c))),
                'cl_cd':       float(np.max(_metric_cl_cd(cl_c,   cd_c))),
                'cl32_cd':     float(np.max(_metric_cl32_cd(cl_c, cd_c))),
                'cl12_cd':     float(np.max(_metric_cl12_cd(cl_c, cd_c))),
                'alpha_stall': float(self.alpha[int(np.argmax(cl_c))]),
            }

        if not raw:
            return {}

        names_list = list(raw)

        def _norm(key):
            vals = np.array([raw[n][key] for n in names_list], dtype=float)
            lo, hi = vals.min(), vals.max()
            if np.isclose(lo, hi):
                return np.ones(len(names_list))
            n = (vals - lo) / (hi - lo)
            return n if self._SCORE_METRICS_HIB[key] else 1.0 - n

        norms = {k: _norm(k) for k in self._SCORE_METRICS_HIB}

        return self._sort_dict({
            name: sum(weights.get(k, 0.0) * norms[k][i] for k in norms)
            for i, name in enumerate(names_list)
        })

    def summary_report(self, airfoil_names, re_target=3e5, scores_dict=None):
        """
        Formatted terminal table. Pass scores_dict to add a Score column.
        Includes all five aerodynamic metrics in the header.
        """
        re_idx = self._re_index(re_target)
        has_score = scores_dict is not None
        sep = '-' * (118 if has_score else 108)

        print(f"\n{sep}")
        print(f" AIRFOIL SUMMARY REPORT  (Re = {re_target:.1e})")
        print(sep)
        hdr = (f"{'Airfoil':<20} | {'Max CL':>7} | {'Min CD':>8} | "
               f"{'CL/CD':>7} | {'CL^1.5/CD':>10} | {'CL^0.5/CD':>10} | "
               f"{'Thick%':>7} | {'Camb%':>6}")
        if has_score:
            hdr += f" | {'Score':>7}"
        print(hdr)
        print(sep)

        for name in airfoil_names:
            try:
                idx = self._airfoil_index(name)
            except KeyError:
                print(f"  {name}: not found"); continue

            cl_c = np.asarray(self.cl[idx][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[idx][re_idx], dtype=float)
            row = (
                f"{name:<20} | "
                f"{np.max(cl_c):>7.4f} | "
                f"{np.min(cd_c):>8.5f} | "
                f"{np.max(_metric_cl_cd(cl_c,   cd_c)):>7.2f} | "
                f"{np.max(_metric_cl32_cd(cl_c, cd_c)):>10.2f} | "
                f"{np.max(_metric_cl12_cd(cl_c, cd_c)):>10.2f} | "
                f"{self.thickness[idx]*100:>7.2f} | "
                f"{self.camber[idx]*100:>6.2f}"
            )
            if has_score:
                row += f" | {scores_dict.get(name, 0.0):>7.4f}"
            print(row)
        print(f"{sep}\n")

    def export_to_excel(self, airfoil_names, re_target,
                        filepath='airfoil_results.xlsx',
                        scores_dict=None):
        """
        Export a formatted summary to Excel.

        Parameters
        ----------
        airfoil_names : list[str]
        re_target     : float
        filepath      : str
        scores_dict   : dict, optional — output of score_airfoils(); adds Score
                        column and sorts the sheet by score descending.

        Returns
        -------
        pandas.DataFrame
        """
        re_idx = self._re_index(re_target)
        rows = []

        for name in airfoil_names:
            try:
                idx = self._airfoil_index(name)
            except KeyError:
                continue

            cl_c = np.asarray(self.cl[idx][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[idx][re_idx], dtype=float)
            cm_c = np.asarray(self.cm[idx][re_idx], dtype=float)

            row = {
                'Airfoil':           name,
                'Max CL':            round(float(np.max(cl_c)),                           4),
                'Min CD':            round(float(np.min(cd_c)),                           5),
                'Min |Cm|':          round(float(np.min(np.abs(cm_c))),                   4),
                'Max CL/CD':         round(float(np.max(_metric_cl_cd(cl_c,   cd_c))),    2),
                'Max CL^1.5/CD':     round(float(np.max(_metric_cl32_cd(cl_c, cd_c))),    2),
                'Max CL^0.5/CD':     round(float(np.max(_metric_cl12_cd(cl_c, cd_c))),    2),
                'Alpha Stall (deg)': float(self.alpha[int(np.argmax(cl_c))]),
                'Thickness (%)':     round(float(self.thickness[idx]) * 100, 2),
                'Camber (%)':        round(float(self.camber[idx])    * 100, 2),
            }
            if scores_dict is not None:
                row['Score'] = round(scores_dict.get(name, 0.0), 4)
            rows.append(row)

        if not rows:
            print("No valid airfoils to export.")
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        if scores_dict is not None and 'Score' in df.columns:
            df = df.sort_values('Score', ascending=False).reset_index(drop=True)

        wb  = Workbook()
        ws  = wb.active
        ws.title = f"Re={re_target:.0e}"

        HDR_FILL = PatternFill('solid', start_color='1F497D')
        HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        SCR_FILL = PatternFill('solid', start_color='E2EFDA')
        ALT_FILL = PatternFill('solid', start_color='F2F2F2')
        THIN     = Side(style='thin', color='BFBFBF')
        BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER   = Alignment(horizontal='center', vertical='center')

        ncols = len(df.columns)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        tc = ws.cell(row=1, column=1,
                     value=f"Airfoil Summary Report  —  Re = {re_target:.2e}")
        tc.font      = Font(bold=True, name='Arial', size=13, color='1F497D')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        # Sub-header
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sc = ws.cell(row=2, column=1,
                     value=("Metrics at Re = {:.2e}  |  {}".format(
                         re_target,
                         "sorted by Score" if scores_dict else "order as supplied")))
        sc.font      = Font(italic=True, name='Arial', size=10, color='595959')
        sc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 16

        # Column headers (row 3)
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=3, column=ci, value=col)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[3].height = 18

        # Data rows
        for ri, row_data in df.iterrows():
            er  = ri + 4
            alt = ALT_FILL if ri % 2 == 1 else PatternFill()
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=er, column=ci, value=val)
                c.font      = Font(bold=(ci == 1), name='Arial', size=10)
                c.alignment = CENTER
                c.border    = BORDER
                c.fill = SCR_FILL if (scores_dict and df.columns[ci-1] == 'Score') else alt

        # Auto column widths
        for ci, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                max(len(str(col)),
                    max(len(str(v)) for v in df.iloc[:, ci-1])) + 4)

        ws.freeze_panes = 'A4'
        wb.save(filepath)
        print(f"Exported -> {filepath}  ({len(df)} airfoils, Re={re_target:.2e})")
        return df

    def plot_shapes(self, airfoil_names, db_dir='airfoilsdb'):
        plt.figure(figsize=(10, 4))
        for name in airfoil_names:
            coords, clean = None, name.replace('.dat', '').strip()
            try:
                af = asb.Airfoil(clean)
                if af.coordinates is not None:
                    coords = af.coordinates
            except Exception:
                pass
            if coords is None:
                fp = os.path.join(db_dir, name)
                if os.path.exists(fp):
                    try:
                        coords = np.loadtxt(fp, skiprows=1)
                    except Exception as e:
                        print(f"Could not parse {fp}: {e}")
            if coords is not None:
                plt.plot(coords[:, 0], coords[:, 1], label=clean)
            else:
                print(f"No coordinates found for {name}")

        plt.title("Airfoil Geometries")
        plt.xlabel("x/c"); plt.ylabel("y/c")
        plt.axis("equal")
        plt.grid(True, linestyle='--', alpha=0.7)
        if plt.gca().get_legend_handles_labels()[1]:
            plt.legend()
        plt.tight_layout(); plt.show()

    def plot_airfoils(self, airfoil_names, re_target=3e5, plot_type='cl_alpha'):
        """
        Performance curve plots.

        plot_type
        ---------
        'cl_alpha'   — CL vs α
        'cm_alpha'   — Cm vs α
        'cl_cd'      — CL/CD vs α
        'cl32_cd'    — CL^1.5/CD vs α  (endurance factor)
        'cl12_cd'    — CL^0.5/CD vs α  (range / min-power factor)
        'drag_polar' — CL vs CD
        """
        re_idx = self._re_index(re_target)

        AXES = {
            'cl_alpha':   ('Angle of Attack (°)', 'CL'),
            'cm_alpha':   ('Angle of Attack (°)', 'Cm'),
            'cl_cd':      ('Angle of Attack (°)', 'CL / CD'),
            'cl32_cd':    ('Angle of Attack (°)', 'CL^1.5 / CD  (Endurance factor)'),
            'cl12_cd':    ('Angle of Attack (°)', 'CL^0.5 / CD  (Range / min-power factor)'),
            'drag_polar': ('CD',                  'CL'),
        }
        if plot_type not in AXES:
            raise ValueError(f"plot_type must be one of {list(AXES)}.")

        plt.figure(figsize=(10, 6))
        for name in airfoil_names:
            try:
                idx = self._airfoil_index(name)
            except KeyError:
                continue
            cl_c = np.asarray(self.cl[idx][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[idx][re_idx], dtype=float)
            cm_c = np.asarray(self.cm[idx][re_idx], dtype=float)

            curve_map = {
                'cl_alpha':   (self.alpha, cl_c),
                'cm_alpha':   (self.alpha, cm_c),
                'cl_cd':      (self.alpha, _metric_cl_cd(cl_c,   cd_c)),
                'cl32_cd':    (self.alpha, _metric_cl32_cd(cl_c, cd_c)),
                'cl12_cd':    (self.alpha, _metric_cl12_cd(cl_c, cd_c)),
                'drag_polar': (cd_c,       cl_c),
            }
            x, y = curve_map[plot_type]
            plt.plot(x, y, label=name)

        plt.xlabel(AXES[plot_type][0]); plt.ylabel(AXES[plot_type][1])
        plt.title(f"{plot_type.replace('_', ' ').title()}  —  Re = {re_target:.1e}")
        plt.grid(True); plt.legend(); plt.tight_layout(); plt.show()

    def plot_operating_envelope(self, airfoil_name, metric='cl_cd'):
        """
        Contour heatmap of Re × α performance for one airfoil.

        metric: 'cl' | 'cd' | 'cm' | 'cl_cd' | 'cl32_cd' | 'cl12_cd'
        """
        try:
            idx = self._airfoil_index(airfoil_name)
        except KeyError as e:
            print(e); return

        cl_data = [np.asarray(self.cl[idx][j], dtype=float) for j in range(len(self.Re))]
        cd_data = [np.asarray(self.cd[idx][j], dtype=float) for j in range(len(self.Re))]

        metric_map = {
            'cl':      np.array(cl_data),
            'cd':      np.array([np.asarray(self.cd[idx][j], dtype=float) for j in range(len(self.Re))]),
            'cm':      np.array([np.asarray(self.cm[idx][j], dtype=float) for j in range(len(self.Re))]),
            'cl_cd':   np.array([_metric_cl_cd(cl_data[j],   cd_data[j]) for j in range(len(self.Re))]),
            'cl32_cd': np.array([_metric_cl32_cd(cl_data[j], cd_data[j]) for j in range(len(self.Re))]),
            'cl12_cd': np.array([_metric_cl12_cd(cl_data[j], cd_data[j]) for j in range(len(self.Re))]),
        }
        if metric not in metric_map:
            raise ValueError(f"metric must be one of {list(metric_map)}.")

        X, Y = np.meshgrid(self.alpha, self.Re)
        plt.figure(figsize=(10, 6))
        cp = plt.contourf(X, Y, metric_map[metric], levels=50, cmap='viridis')
        plt.colorbar(cp, label=metric)
        plt.yscale('log')
        plt.title(f"{airfoil_name} — Operating Envelope ({metric})")
        plt.xlabel("Angle of Attack (°)"); plt.ylabel("Reynolds Number")
        plt.grid(True, alpha=0.3); plt.tight_layout(); plt.show()

    def plot_pareto_scatter(self, metric_x_dict, metric_y_dict,
                            x_label="Maximise X", y_label="Minimise Y"):
        """Pareto scatter: X is maximised, Y is minimised."""
        names = list(set(metric_x_dict).intersection(metric_y_dict))
        if not names:
            print("No common airfoils."); return

        df = (pd.DataFrame({'Name': names,
                            'X': [metric_x_dict[n] for n in names],
                            'Y': [metric_y_dict[n] for n in names]})
              .sort_values('X', ascending=False).reset_index(drop=True))

        min_y, pidx = float('inf'), []
        for i, row in df.iterrows():
            if row['Y'] < min_y:
                pidx.append(i); min_y = row['Y']
        df['Is_Pareto'] = False
        df.loc[pidx, 'Is_Pareto'] = True

        plt.figure(figsize=(12, 7))
        sns.scatterplot(data=df[~df['Is_Pareto']], x='X', y='Y',
                        color='lightgray', alpha=0.6, s=50, label='Dominated')
        sns.scatterplot(data=df[df['Is_Pareto']], x='X', y='Y',
                        color='red', s=120, edgecolor='black', zorder=5,
                        label='Pareto Frontier')
        front = df[df['Is_Pareto']].sort_values('X')
        plt.plot(front['X'], front['Y'], 'r--', alpha=0.5, zorder=4)
        for _, row in front.iterrows():
            plt.annotate(row['Name'], (row['X'], row['Y']),
                         xytext=(8, 5), textcoords='offset points',
                         fontsize=9, fontweight='bold')

        plt.title(f"Trade-off: {x_label} vs {y_label}", fontsize=13, pad=12)
        plt.xlabel(x_label, fontsize=11, fontweight='bold')
        plt.ylabel(y_label, fontsize=11, fontweight='bold')
        plt.grid(True, linestyle=':', alpha=0.7)
        plt.legend(frameon=True, shadow=True)
        plt.tight_layout(); plt.show()

    def plot_radar_chart(self, airfoil_names, re_target=5e5):
        """
        Normalised radar chart comparing candidates across all five aerodynamic
        metrics plus geometry. All axes in [0, 1].
        """
        re_idx = self._re_index(re_target)
        categories = ['Max CL', 'Low CD', 'CL/CD', 'CL^1.5/CD', 'CL^0.5/CD',
                      'Thickness', 'Camber']

        raw = {}
        for name in airfoil_names:
            try:
                idx = self._airfoil_index(name)
            except KeyError:
                continue
            cl_c = np.asarray(self.cl[idx][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[idx][re_idx], dtype=float)
            raw[name] = {
                'cl':      float(np.max(cl_c)),
                'cd':      float(np.min(cd_c)),
                'cl_cd':   float(np.max(_metric_cl_cd(cl_c,   cd_c))),
                'cl32_cd': float(np.max(_metric_cl32_cd(cl_c, cd_c))),
                'cl12_cd': float(np.max(_metric_cl12_cd(cl_c, cd_c))),
                'thick':   float(self.thickness[idx]),
                'camber':  float(self.camber[idx]),
            }

        if not raw:
            return

        keys = ['cl', 'cd', 'cl_cd', 'cl32_cd', 'cl12_cd', 'thick', 'camber']
        hibs = [True, False, True, True, True, True, True]

        def _n(key, hib):
            vals = np.array([raw[n][key] for n in raw])
            lo, hi = vals.min(), vals.max()
            if np.isclose(lo, hi): return np.ones(len(vals))
            n = (vals - lo) / (hi - lo)
            return n if hib else 1.0 - n

        norms = {k: _n(k, h) for k, h in zip(keys, hibs)}

        fig = go.Figure()
        for i, name in enumerate(raw):
            vals = [norms[k][i] for k in keys] + [norms[keys[0]][i]]
            fig.add_trace(go.Scatterpolar(
                r=vals, theta=categories + [categories[0]],
                fill='toself', name=name))

        fig.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 1])),
            showlegend=True,
            title=f"Multi-Metric Radar  (normalised)  —  Re = {re_target:.1e}")
        fig.show()

    def plot_multi_re_drag_polar(self, airfoil_name):
        """Drag polar across all Re for a single airfoil."""
        try:
            idx = self._airfoil_index(airfoil_name)
        except KeyError as e:
            print(e); return

        plt.figure(figsize=(10, 8))
        colors = plt.cm.jet(np.linspace(0, 1, len(self.Re)))
        for i, re_val in enumerate(self.Re):
            plt.plot(self.cd[idx][i], self.cl[idx][i],
                     label=f'Re = {re_val:.1e}', color=colors[i])
        plt.title(f"{airfoil_name} — Drag Polar Across All Re")
        plt.xlabel("CD"); plt.ylabel("CL")
        plt.xlim(0, np.percentile(self.cd[idx], 80))
        plt.grid(True)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.tight_layout(); plt.show()

    def plot_parallel_coordinates(self, re_target=5e5):
        """Parallel coordinates for the full database."""
        re_idx = self._re_index(re_target)

        rows = []
        for i in range(len(self.names)):
            cl_c = np.asarray(self.cl[i][re_idx], dtype=float)
            cd_c = np.asarray(self.cd[i][re_idx], dtype=float)
            rows.append({
                'Thickness':   float(self.thickness[i]),
                'Camber':      float(self.camber[i]),
                'Max_CL':      float(np.max(cl_c)),
                'Min_CD':      float(np.min(cd_c)),
                'Max_CL_CD':   float(np.max(_metric_cl_cd(cl_c,   cd_c))),
                'Max_CL32_CD': float(np.max(_metric_cl32_cd(cl_c, cd_c))),
                'Max_CL12_CD': float(np.max(_metric_cl12_cd(cl_c, cd_c))),
            })
        df_plot = pd.DataFrame(rows)

        fig = px.parallel_coordinates(
            df_plot, color='Max_CL_CD',
            labels={
                'Thickness':   'Thickness',
                'Camber':      'Camber',
                'Max_CL':      f'Max CL  (Re={re_target:.0e})',
                'Min_CD':      f'Min CD  (Re={re_target:.0e})',
                'Max_CL_CD':   f'Max L/D (Re={re_target:.0e})',
                'Max_CL32_CD': f'Max L^1.5/D (Re={re_target:.0e})',
                'Max_CL12_CD': f'Max L^0.5/D (Re={re_target:.0e})',
            },
            color_continuous_scale=px.colors.diverging.Tealrose,
            color_continuous_midpoint=df_plot['Max_CL_CD'].median(),
        )
        fig.update_layout(title="Airfoil Database — Parallel Coordinates")
        fig.show()